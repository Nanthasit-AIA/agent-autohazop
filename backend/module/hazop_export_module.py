"""Styled HAZOP/LOPA workbook export for the v1.5 20-column schema.

The rendered layout follows the SCG worksheet format (GW / DEVIATION / CAUSES /
CONSEQUENCES / S / L / R / IPLs / Meet IPL criteria / SAFEGUARDS / S / L / R /
RECOMMENDATIONS / COMMENTS). Node, Design Intention and Parameter are not table
columns there - they belong in the metadata block above the table, which is how
the client worksheet presents them.
"""
from pathlib import Path
from typing import Any, Dict, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from module.prompt.hzp_promptt import HAZOP_LOPA_HEADERS_20

# Rendered column -> source column in HAZOP_LOPA_HEADERS_20.
RENDER_MAP = [
    ("A", "GW", "Guide Word", 10),
    ("B", "DEVIATION", "Deviation", 16),
    ("C", "CAUSES", "Cause", 42),
    ("D", "CONSEQUENCES", "Consequence", 46),
    ("E", "S", "Initial Severity", 5),
    ("F", "L", "Initial Likelihood", 5),
    ("G", "R", "Initial Risk Ranking", 6),
    ("H", "IPLs", "IPLs", 34),
    ("I", "Independent", "IPL Independent", 5),
    ("J", "Effective", "IPL Effective", 5),
    ("K", "Auditable", "IPL Auditable", 5),
    ("L", "SAFEGUARDS", "Safeguards", 34),
    ("M", "S", "Final Severity", 5),
    ("N", "L", "Final Likelihood", 5),
    ("O", "R", "Final Risk Ranking", 6),
    ("P", "RECOMMENDATIONS", "Recommendations", 38),
    ("Q", "COMMENTS/ACTIONS/OPPORTUNITY FOR IMPROVEMENTS", "Comments_Actions", 34),
]

HEADER_ROW_TOP = 8
HEADER_ROW_SUB = 9
DATA_START_ROW = 10

_THIN = Side(style="thin", color="B4C6E7")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def normalize_hazop_dataframe(df: pd.DataFrame | None) -> pd.DataFrame:
    """Return df reindexed onto the canonical 20 columns, without dropping rows."""
    if df is None or df.empty:
        return pd.DataFrame(columns=HAZOP_LOPA_HEADERS_20)
    out = df.copy()
    out = out.loc[:, ~out.columns.astype(str).str.startswith("Unnamed")]
    return out.reindex(columns=HAZOP_LOPA_HEADERS_20).fillna("")


def _text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _first_non_empty(df: pd.DataFrame, column: str) -> str:
    if column not in df.columns:
        return ""
    for value in df[column]:
        text = _text(value)
        if text:
            return text
    return ""


def _unique_join(df: pd.DataFrame, column: str) -> str:
    if column not in df.columns:
        return ""
    seen: list[str] = []
    for value in df[column]:
        text = _text(value)
        if text and text not in seen:
            seen.append(text)
    return ", ".join(seen)


def _metadata_summary(df: pd.DataFrame) -> Dict[str, str]:
    return {
        "node": _first_non_empty(df, "Node"),
        "design_intention": _first_non_empty(df, "Design Intention"),
        "parameter": _unique_join(df, "Parameter"),
    }


def _write_metadata(ws, metadata: Dict[str, str]) -> None:
    ws.merge_cells("A1:Q1")
    ws["A1"] = "INTEGRATED HAZOP/LOPA WORKSHEET"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    label_font = Font(name="Arial", size=10, bold=True)
    value_font = Font(name="Arial", size=10)

    ws.merge_cells("B3:C3")
    ws.merge_cells("E3:G3")
    ws.merge_cells("J3:Q3")
    ws.merge_cells("B4:C5")
    ws.merge_cells("E4:K5")
    ws.merge_cells("B6:Q6")

    # Company / Facility / Drawings are not carried in the generated rows; the
    # labels are written so the engineer can complete them in the workbook.
    for cell, label in {
        "A3": "Company", "D3": "Facility", "I3": "Drawings",
        "A4": "Node", "D4": "Design intention", "A6": "Parameter",
    }.items():
        ws[cell] = label
        ws[cell].font = label_font
        ws[cell].alignment = Alignment(vertical="top")

    for cell, value in {
        "B4": metadata["node"],
        "E4": metadata["design_intention"],
        "B6": metadata["parameter"],
    }.items():
        ws[cell] = value
        ws[cell].font = value_font
        ws[cell].alignment = Alignment(vertical="top", wrap_text=True)

    ws.merge_cells("A7:Q7")
    ws["A7"] = (
        "S: Severity of consequence    L: Likelihood    R: Risk ranking.    "
        "IPL criteria columns record engineer verification of independence, effectiveness and auditability."
    )
    ws["A7"].font = Font(name="Arial", size=9, italic=True, color="404040")
    ws["A7"].alignment = Alignment(wrap_text=True, vertical="center")


def _write_table_header(ws) -> None:
    header_font = Font(name="Arial", size=9, bold=True)
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    for col, label, _source, _width in RENDER_MAP:
        if col in {"I", "J", "K"}:
            continue
        ws.merge_cells(f"{col}{HEADER_ROW_TOP}:{col}{HEADER_ROW_SUB}")
        cell = ws[f"{col}{HEADER_ROW_TOP}"]
        cell.value = label
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    ws.merge_cells(f"I{HEADER_ROW_TOP}:K{HEADER_ROW_TOP}")
    group = ws[f"I{HEADER_ROW_TOP}"]
    group.value = "Meet IPL criteria?"
    group.font = header_font
    group.fill = header_fill
    group.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    group.border = _BORDER

    for col, label, _source, _width in RENDER_MAP:
        if col not in {"I", "J", "K"}:
            continue
        cell = ws[f"{col}{HEADER_ROW_SUB}"]
        cell.value = label
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)
        cell.border = _BORDER


def _write_data_rows(ws, df: pd.DataFrame) -> int:
    row_idx = DATA_START_ROW
    for _, record in df.iterrows():
        for col, _label, source, _width in RENDER_MAP:
            cell = ws[f"{col}{row_idx}"]
            cell.value = _text(record.get(source, ""))
            centered = col in {"E", "F", "G", "I", "J", "K", "M", "N", "O"}
            cell.alignment = Alignment(
                horizontal="center" if centered else "left",
                vertical="top",
                wrap_text=not centered,
            )
            cell.font = Font(name="Arial", size=9)
            cell.border = _BORDER
        row_idx += 1
    return row_idx - 1


def _style_worksheet(ws, last_row: int) -> None:
    for col, _label, _source, width in RENDER_MAP:
        ws.column_dimensions[col].width = width
    ws.freeze_panes = f"A{DATA_START_ROW}"
    if last_row >= DATA_START_ROW:
        ws.auto_filter.ref = f"A{HEADER_ROW_SUB}:Q{last_row}"


def _write_raw_data_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Full 20-column source data, so nothing is lost by the rendered layout."""
    ws = wb.create_sheet("Raw Data")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")

    for col_idx, header in enumerate(HAZOP_LOPA_HEADERS_20, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    for row_idx, (_, record) in enumerate(df.iterrows(), start=2):
        for col_idx, header in enumerate(HAZOP_LOPA_HEADERS_20, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_text(record.get(header, "")))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _BORDER

    for col_idx in range(1, len(HAZOP_LOPA_HEADERS_20) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.freeze_panes = "A2"


def _preview_labels() -> list[str]:
    """Rendered labels made unique. S/L/R appear twice in the worksheet (initial and
    final); the Excel layout distinguishes them by position, but a DataFrame needs
    distinct names or the later column silently overwrites the earlier one.
    """
    counts: Dict[str, int] = {}
    for _col, label, _source, _width in RENDER_MAP:
        counts[label] = counts.get(label, 0) + 1

    labels: list[str] = []
    for _col, label, source, _width in RENDER_MAP:
        if counts[label] > 1:
            labels.append(f"{label} ({source.split()[0].lower()})")
        else:
            labels.append(label)
    return labels


def hazop_lopa_preview_dataframe(df: pd.DataFrame | None) -> pd.DataFrame:
    """Rendered-layout view of the rows, for the UI result preview."""
    labels = _preview_labels()
    normalized = normalize_hazop_dataframe(df)
    if normalized.empty:
        return pd.DataFrame(columns=labels)
    out = pd.DataFrame()
    for label, (_col, _rendered, source, _width) in zip(labels, RENDER_MAP):
        out[label] = normalized[source].map(_text)
    return out


def write_hazop_lopa_workbook(df: pd.DataFrame, output_path: str) -> None:
    normalized = normalize_hazop_dataframe(df)

    wb = Workbook()
    ws = wb.active
    ws.title = "HAZOP_LOPA"

    _write_metadata(ws, _metadata_summary(normalized))
    _write_table_header(ws)
    last_row = _write_data_rows(ws, normalized)
    _style_worksheet(ws, last_row)
    _write_raw_data_sheet(wb, normalized)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
