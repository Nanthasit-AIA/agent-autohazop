"""Engineer review round-trip: P&ID JSON -> Excel -> edited Excel -> P&ID JSON.

The sheet layout is defined once in SECTIONS and drives both directions, so export
and import cannot drift apart. Every field of the v1.5 PIDResponse schema is
represented, including node_define, intention, the extended Connection fields and
the line-level sheets - a round-trip must not silently drop them.
"""
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from decorators import logger
from module.schema_json import PIDResponse

LIST_SEP = "; "

# (sheet, json key, [(column header, field name, kind)]) where kind is "text" or "list".
SECTIONS: list[tuple[str, str, list[tuple[str, str, str]]]] = [
    ("Node Define", "node_define", [
        ("node_id", "node_id", "text"),
        ("node_name", "node_name", "text"),
        ("description", "description", "text"),
    ]),
    ("Equipment", "equipment", [
        ("id", "id", "text"), ("name", "name", "text"),
        ("type", "type", "text"), ("context", "context", "text"),
    ]),
    ("Valves", "valves", [
        ("id", "id", "text"), ("type", "type", "text"),
        ("location", "location", "text"), ("context", "context", "text"),
    ]),
    ("Instruments", "instruments", [
        ("id", "id", "text"), ("function", "function", "text"),
        ("location", "location", "text"), ("context", "context", "text"),
    ]),
    ("Utility Lines", "utility_lines", [
        ("utility_type", "utility_type", "text"),
        ("valves", "valves", "list"),
        ("flow_direction", "flow_direction", "text"),
        ("context", "context", "text"),
    ]),
    ("Connections", "connections", [
        ("line_id", "line_id", "text"),
        ("node", "node", "text"),
        ("node_boundary", "node_boundary", "text"),
        ("from_id", "from_id", "text"),
        ("to_id", "to_id", "text"),
        ("included_equipment", "included_equipment", "list"),
        ("included_lines", "included_lines", "list"),
        ("valves", "valves", "list"),
        ("instruments", "instruments", "list"),
        ("flow_direction", "flow_direction", "text"),
        ("context", "context", "text"),
    ]),
    ("Line Connections", "line_level_connections", [
        ("line_id", "line_id", "text"),
        ("from_id", "from_id", "text"),
        ("to_id", "to_id", "text"),
        ("valves", "valves", "list"),
        ("instruments", "instruments", "list"),
        ("flow_direction", "flow_direction", "text"),
        ("context", "context", "text"),
    ]),
]

# line_level_details nests line_id_parse and nominal_size, so it is flattened.
LINE_DETAIL_SHEET = "Line Details"
LINE_DETAIL_COLUMNS = [
    "line_id", "raw_line_label", "area_or_unit", "service_code",
    "line_sequence_number", "piping_class", "insulation_or_design_code",
    "nominal_size_value", "nominal_size_unit",
]

PROCESS_FIELDS = ["process_description", "intention"]


def _pid_root(data: Any) -> dict[str, Any]:
    if isinstance(data, dict) and isinstance(data.get("pid_data"), dict):
        return data["pid_data"]
    return data if isinstance(data, dict) else {}


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _join(value: Any) -> str:
    if isinstance(value, list):
        return LIST_SEP.join(_text(v) for v in value if _text(v))
    return _text(value)


def _split(value: Any) -> list[str]:
    raw = _text(value)
    if not raw:
        return []
    parts = raw.replace("\n", ";").replace(",", ";").split(";")
    return [p.strip() for p in parts if p.strip()]


def _style_sheet(ws, columns: list[str], row_count: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    for idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = 16 if len(header) < 14 else 26
    for r in range(2, row_count + 2):
        for c in range(1, len(columns) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def export_pid_review_excel(data: Any, out_dir: str | Path, name: str) -> Path:
    pid = _pid_root(data)
    wb = Workbook()

    ws = wb.active
    ws.title = "Process"
    rows = [(f, _text(pid.get(f, ""))) for f in PROCESS_FIELDS]
    for i, (field, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=field)
        ws.cell(row=i, column=2, value=value)
    _style_sheet(ws, ["field", "value"], len(rows))
    ws.column_dimensions["B"].width = 90

    io = wb.create_sheet("Inputs Outputs")
    io_rows = [("system_input", _text(v)) for v in pid.get("system_inputs") or []]
    io_rows += [("system_output", _text(v)) for v in pid.get("system_outputs") or []]
    for i, (kind, nm) in enumerate(io_rows, start=2):
        io.cell(row=i, column=1, value=kind)
        io.cell(row=i, column=2, value=nm)
    _style_sheet(io, ["kind", "name"], len(io_rows))

    for sheet_name, key, columns in SECTIONS:
        sh = wb.create_sheet(sheet_name)
        records = pid.get(key) or []
        for r, record in enumerate(records, start=2):
            record = record if isinstance(record, dict) else {}
            for c, (_header, field, kind) in enumerate(columns, start=1):
                value = record.get(field)
                sh.cell(row=r, column=c, value=_join(value) if kind == "list" else _text(value))
        _style_sheet(sh, [h for h, _f, _k in columns], len(records))

    ld = wb.create_sheet(LINE_DETAIL_SHEET)
    details = pid.get("line_level_details") or []
    for r, record in enumerate(details, start=2):
        record = record if isinstance(record, dict) else {}
        parse = record.get("line_id_parse") or {}
        size = parse.get("nominal_size") or {}
        values = [
            _text(record.get("line_id")), _text(record.get("raw_line_label")),
            _text(parse.get("area_or_unit")), _text(parse.get("service_code")),
            _text(parse.get("line_sequence_number")), _text(parse.get("piping_class")),
            _text(parse.get("insulation_or_design_code")),
            _text(size.get("value")), _text(size.get("unit")),
        ]
        for c, value in enumerate(values, start=1):
            ld.cell(row=r, column=c, value=value)
    _style_sheet(ld, LINE_DETAIL_COLUMNS, len(details))

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{Path(name).stem}_pid_review.xlsx"
    wb.save(path)
    return path


def _read_rows(wb, sheet_name: str, columns: list[str]) -> list[dict[str, str]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    header = [_text(c.value) for c in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw is None or all(_text(v) == "" for v in raw):
            continue
        record = {}
        for col in columns:
            record[col] = _text(raw[header.index(col)]) if col in header and header.index(col) < len(raw) else ""
        rows.append(record)
    return rows


def _check_duplicates(rows: list[dict[str, str]], key: str, label: str, errors: list[str]) -> None:
    seen, dupes = set(), set()
    for row in rows:
        value = row.get(key, "")
        if not value:
            continue
        if value in seen:
            dupes.add(value)
        seen.add(value)
    for value in sorted(dupes):
        errors.append(f"{label}: duplicate {key} '{value}'")


def import_pid_review_excel(path: str | Path) -> tuple[dict[str, Any], list[str], list[str]]:
    wb = load_workbook(path, data_only=True)
    errors: list[str] = []
    warnings: list[str] = []

    process = {r["field"]: r["value"] for r in _read_rows(wb, "Process", ["field", "value"])}
    pid: dict[str, Any] = {f: process.get(f, "") for f in PROCESS_FIELDS}

    io_rows = _read_rows(wb, "Inputs Outputs", ["kind", "name"])
    pid["system_inputs"] = [r["name"] for r in io_rows if r["kind"].lower().startswith("system_input")]
    pid["system_outputs"] = [r["name"] for r in io_rows if r["kind"].lower().startswith("system_output")]

    for sheet_name, key, columns in SECTIONS:
        headers = [h for h, _f, _k in columns]
        rows = _read_rows(wb, sheet_name, headers)
        records = []
        for row in rows:
            record: dict[str, Any] = {}
            for header, field, kind in columns:
                record[field] = _split(row.get(header, "")) if kind == "list" else row.get(header, "")
            records.append(record)
        pid[key] = records
        if key in {"equipment", "valves", "instruments"}:
            _check_duplicates(rows, "id", sheet_name, errors)
        if key == "node_define":
            _check_duplicates(rows, "node_id", sheet_name, errors)

    details = []
    for row in _read_rows(wb, LINE_DETAIL_SHEET, LINE_DETAIL_COLUMNS):
        size = None
        if row["nominal_size_value"]:
            try:
                size = {"value": float(row["nominal_size_value"]), "unit": row["nominal_size_unit"] or ""}
            except ValueError:
                warnings.append(
                    f"{LINE_DETAIL_SHEET}: line '{row['line_id']}' has a non-numeric "
                    f"nominal_size_value '{row['nominal_size_value']}'; size dropped"
                )
        parse = {
            "area_or_unit": row["area_or_unit"] or None,
            "service_code": row["service_code"] or None,
            "line_sequence_number": row["line_sequence_number"] or None,
            "piping_class": row["piping_class"] or None,
            "insulation_or_design_code": row["insulation_or_design_code"] or None,
            "nominal_size": size,
        }
        details.append({
            "line_id": row["line_id"],
            "raw_line_label": row["raw_line_label"] or None,
            "line_id_parse": parse if any(v for v in parse.values()) else None,
        })
    pid["line_level_details"] = details

    if not errors:
        try:
            PIDResponse(**pid)
        except Exception as exc:
            errors.append(f"Schema validation failed: {exc}")

    return pid, errors, warnings


def save_reviewed_pid_json(pid_data: dict[str, Any], out_dir: str | Path, name: str, source_file: str) -> Path:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{Path(name).stem}.json"
    combined = {
        "pid_data": pid_data,
        "metadata": {
            "source": "Engineer reviewed Excel",
            "source_file": source_file,
            "reviewed_at": datetime.now().isoformat(timespec="seconds"),
        },
    }
    path.write_text(json.dumps(combined, indent=2, ensure_ascii=False), encoding="utf-8")
    logger.info(f"Saved reviewed P&ID JSON: {path}")
    return path
